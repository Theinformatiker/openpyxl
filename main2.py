import openpyxl

# Pfad zur Original-Excel-Datei
original_excel_file_path = "C://Users//vkkor//Documents//Udemy2024//ExcelPy//Wetterinformationen.xlsx"

# Original-Arbeitsmappe öffnen
original_workbook = openpyxl.load_workbook(original_excel_file_path)

# Original-Arbeitsblatt auswählen
original_worksheet = original_workbook["Wetterinformationen"]

# Anzahl der Spalten im Original-Arbeitsblatt
num_columns = original_worksheet.max_column

# Durchlaufe die Zeilen im Original-Arbeitsblatt
for i in range(2, original_worksheet.max_row + 1):
    # Extrahiere den Städtenamen aus der Spalte A
    city = original_worksheet.cell(row=i, column=1).value

    # Neue Arbeitsmappe erstellen
    new_workbook = openpyxl.Workbook()
    new_worksheet = new_workbook.active

    # Kopiere die Überschriften ins neue Arbeitsblatt
    for j in range(1, num_columns + 1):
        new_worksheet.cell(row=1, column=j, value=original_worksheet.cell(row=1, column=j).value)

    # Kopiere die aktuelle Zeile ins neue Arbeitsblatt
    for j in range(1, num_columns + 1):
        new_worksheet.cell(row=2, column=j, value=original_worksheet.cell(row=i, column=j).value)

    # Speichere die neue Arbeitsmappe
    new_excel_file_path = f"C://Users//vkkor//Documents//Udemy2024//ExcelPy//{city}.xlsx"
    new_workbook.save(new_excel_file_path)

    print(f"Arbeitsmappe für Zeile {i} mit der Stadt {city} wurde gespeichert: {new_excel_file_path}")

# Original-Arbeitsmappe schließen
original_workbook.close()
